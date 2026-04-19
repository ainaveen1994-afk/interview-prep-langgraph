# =============================================================================

# FINAL: Resume vs JD Matcher (Hybrid ATS + Excel + LangGraph)

# =============================================================================

import sys
import operator
import json
import os
from typing import Annotated
from datetime import datetime

from dotenv import load_dotenv
from pydantic import BaseModel
from langchain_openai import ChatOpenAI
from langgraph.graph import StateGraph, START, END

from PyPDF2 import PdfReader
import docx
from openpyxl import Workbook, load_workbook

sys.stdout.reconfigure(encoding="utf-8")
load_dotenv()

llm = ChatOpenAI(model="gpt-4o-mini", temperature=0.3)

# =============================================================================

# STATE

# =============================================================================

class ResumeMatchState(BaseModel):
    job_description: dict = {}
    resume_text: str = ""
    resume_profile: dict = {}

    match_score: float = 0.0

    missing_skills: dict = {}

    decision: str = ""
    final_report: str = ""

    messages: Annotated[list, operator.add] = []


# =============================================================================

# HELPERS

# =============================================================================

def extract_text(file_path: str) -> str:
    try:
        if file_path.endswith(".pdf"):
            reader = PdfReader(file_path)
            return " ".join([p.extract_text() or "" for p in reader.pages])

        elif file_path.endswith(".docx"):
            doc = docx.Document(file_path)
            return "\n".join([p.text for p in doc.paragraphs])

        else:
            return ""

    except Exception as e:
        return f"Error reading file: {e}"


def extract_jd_json(file_path):
    try:
        # ✅ NEW: handle JSON directly
        if file_path.endswith(".json"):
            with open(file_path, "r", encoding="utf-8") as f:
                return json.load(f)

        # existing logic for docx/pdf
        text = extract_text(file_path)
        start = text.index("{")
        end = text.rindex("}") + 1
        return json.loads(text[start:end])

    except Exception as e:
        print(f"❌ Failed to extract JD JSON: {e}")
        return {}



def save_to_excel(result, resume_path, excel_path=None):
    file_name = os.path.basename(resume_path)


    if not excel_path:
        today = datetime.now().strftime("%Y-%m-%d")
        excel_path = f"resume_results_{today}.xlsx"

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([
            "Candidate",
            "File",
            "Score",
            "Decision",
            "Missing Core Skills",
            "Missing AI Skills",
            "Missing Tools",
            "Date"
        ])

    

    
    missing = result["missing_skills"]

    file_name = os.path.basename(resume_path)

    candidate_name = result.get("resume_profile", {}).get("name", "")

    if not candidate_name:
        candidate_name = file_name.replace(".docx","").replace(".pdf","")

    ws.append([
        candidate_name,
        file_name,
        result["match_score"],
        result["decision"],
        ", ".join(missing.get("core", [])),
        ", ".join(missing.get("ai", [])),
        ", ".join(missing.get("tools", [])),
        datetime.now().strftime("%Y-%m-%d %H:%M")
    ])


    wb.save(excel_path)

    print(f"✅ Saved to Excel: {excel_path}")

# =============================================================================

# NODES

# =============================================================================

def load_documents(state: ResumeMatchState) -> dict:
    jd_json = state.job_description   # ✅ already dict
    resume_text = extract_text(state.resume_text)

    return {
        "job_description": jd_json,
        "resume_text": resume_text,
        "messages": ["[load_documents] Loaded"]
    }

def extract_resume_profile(state: ResumeMatchState) -> dict:
    response = llm.invoke(
       
f"""
Extract candidate details from resume.

Return ONLY JSON:

{{
  "name": "Full Name",
  "skills": [],
  "experience": number
}}

Rules:
- Extract full name from top of resume
- Do NOT guess
- Include all technical skills

RESUME:
{state.resume_text}
"""
)


    content = response.content.strip()

    # ✅ Debug (optional)
    print("LLM RAW OUTPUT:", content)

    try:
        # ✅ Handle markdown like ```json ... ```
        if "```" in content:
            content = content.split("```")[1]
            content = content.replace("json", "").strip()

        data = json.loads(content)

    except Exception as e:
        print("❌ JSON parsing failed:", e)

        # ✅ fallback (prevents crash)
        data = {
            "skills": [],
            "experience": 0
        }

    # ✅ Safe cleanup
    data["skills"] = list(set([s.strip() for s in data.get("skills", [])]))
    data["name"] = data.get("name", "").strip()


    return {
        "resume_profile": data,
        "messages": ["[resume_profile] extracted"]
    }
def normalize(skills):
    return set(s.lower().strip() for s in skills)

def skill_match(jd_skills, resume_skills):
    matched = set()

    for jd in jd_skills:
        for res in resume_skills:
            if jd in res or res in jd:
                matched.add(jd)

    return matched

def compute_score(state: ResumeMatchState) -> dict:
    jd = state.job_description
    resume = state.resume_profile

    # ✅ Normalize skills
    jd_skills = normalize(jd.get("must_have_skills", []))
    ai_skills = normalize(jd.get("ai_automation_skills", []))
    tools = normalize(jd.get("tools_technologies", []))
    resume_skills = normalize(resume.get("skills", []))

    # ✅ Debug (keep for now)
    print("JD Skills:", jd_skills)
    print("Resume Skills:", resume_skills)

    # ✅ Matching
    matched_core = skill_match(jd_skills, resume_skills)
    matched_ai = skill_match(ai_skills, resume_skills)
    matched_tools = skill_match(tools, resume_skills)

    # ✅ Missing skills
    missing_core = list(jd_skills - matched_core)
    missing_ai = list(ai_skills - matched_ai)
    missing_tools = list(tools - matched_tools)

    # ✅ Scores
    core_score = (len(matched_core) / len(jd_skills)) * 100 if jd_skills else 0
    ai_score = (len(matched_ai) / len(ai_skills)) * 100 if ai_skills else 0
    tools_score = (len(matched_tools) / len(tools)) * 100 if tools else 0

    # ✅ Experience score
    jd_exp = jd.get("experience_required", 1)
    res_exp = resume.get("experience", 0)
    exp_score = min(res_exp / jd_exp, 1) * 100

    # ✅ Final weighted score
    final_score = round(
        core_score * 0.5 +
        ai_score * 0.2 +
        tools_score * 0.1 +
        exp_score * 0.2,
        2
    )

    return {
        "match_score": final_score,
        "missing_skills": {
            "core": missing_core,
            "ai": missing_ai,
            "tools": missing_tools
        },
        "resume_profile": resume,  # ✅ needed for candidate name in Excel
        "messages": [f"[score] {final_score}"]
    }





def final_decision(state: ResumeMatchState) -> dict:
    if state.missing_skills:
        decision = "REJECTED"
    elif state.match_score >= 75:
        decision = "SHORTLISTED"
    elif state.match_score >= 60:
        decision = "CONSIDER"
    else:
        decision = "REJECTED"

    reason = llm.invoke(
        f"Explain in 2 lines why candidate is {decision}. "
        f"Score: {state.match_score}, Missing: {state.missing_skills}"
    ).content

    report = f"""


    ================= MATCH REPORT =================

    Score        : {state.match_score}%
    Decision     : {decision}

    Missing Skills:
    {chr(10).join([f"✘ {m}" for m in state.missing_skills])}

    Reason:
    {reason}

    ================================================
    """


    return {
        "decision": decision,
        "final_report": report,
        "messages": [f"[decision] {decision}"]
    }


# =============================================================================

# GRAPH

# =============================================================================

graph = StateGraph(ResumeMatchState)

graph.add_node("load", load_documents)
graph.add_node("resume_profile", extract_resume_profile)
graph.add_node("score", compute_score)
graph.add_node("decision", final_decision)

graph.add_edge(START, "load")
graph.add_edge("load", "resume_profile")
graph.add_edge("resume_profile", "score")
graph.add_edge("score", "decision")
graph.add_edge("decision", END)

app = graph.compile()

# =============================================================================

# RUNNER

# =============================================================================
def run_resume_match(jd_path, resume_path, excel_path=None):
    if not os.path.exists(jd_path):
        print("❌ JD file not found")
        return

    if not os.path.exists(resume_path):
        print("❌ Resume file not found")
        return

    # ✅ LOAD JSON HERE
    jd_data = extract_jd_json(jd_path)

    result = app.invoke({
        "job_description": jd_data,   # ✅ now dict
        "resume_text": resume_path,
        "messages": []
    })

    print(result["final_report"])

    save_to_excel(result, resume_path, excel_path)

    return result



# =============================================================================

# MAIN

# =============================================================================

if __name__ == "__main__":
    print("\n=== RESUME MATCHING SYSTEM ===")


    jd = input("Enter JD file path:\n").strip().strip('"')
    resume = input("Enter Resume file path:\n").strip().strip('"')

    excel_path = input("Enter Excel file path (or press Enter for auto):\n").strip()
    if excel_path == "":
        excel_path = None

    run_resume_match(jd, resume, excel_path)

